import csv
import json
import os
import random
import string
import tempfile
from collections import defaultdict
from io import BytesIO

import pandas as pd
import requests
from django.conf import settings as django_settings
from django.core.mail import send_mail
from django.db.models import Avg, StdDev
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from employees.models import Employee
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework import generics, status
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import AccessCode, Answer, Dimension, Question, SurveyResponse
from .serializers import (
    AccessCodeSerializer,
    DimensionSerializer,
    SurveyResponseSerializer,
)


class DimensionListAPIView(generics.ListAPIView):
    """List all dimensions with their associated questions."""

    queryset = Dimension.objects.prefetch_related("questions").all()
    serializer_class = DimensionSerializer
    permission_classes = [AllowAny]


class SurveyResponseCreateAPIView(generics.CreateAPIView):
    """Create a new survey response, associating it with an employee using an access code."""

    queryset = SurveyResponse.objects.all()
    serializer_class = SurveyResponseSerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        """Save the survey response and associate it with an employee using an access code."""
        access_code = self.request.data.get("access_code")
        print("Access Code:", access_code)

        try:
            # Verificar si el código de acceso existe y no ha sido utilizado
            access_code_obj = AccessCode.objects.get(code=access_code)

        except AccessCode.DoesNotExist:
            raise ValidationError("Código de acceso no válido.")

        # Obtener el empleado asociado al código de acceso
        employee = access_code_obj.employee

        # Obtener las respuestas de la encuesta
        answers_data = self.request.data.get("answers")

        # Crear la respuesta de la encuesta y asociarla al empleado
        survey_response = serializer.save(employee=employee)

        # Crear las respuestas individuales
        answers = []
        for answer_data in answers_data:
            question = Question.objects.get(id=answer_data["question"])
            answers.append(
                Answer(
                    response=survey_response,
                    question=question,
                    value=answer_data["value"],
                )
            )

        # Guardar las respuestas de manera eficiente
        Answer.objects.bulk_create(answers)


def generate_code(length=8):
    """Generate a random alphanumeric code of specified length."""
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=length))


class GenerateAccessCodeView(APIView):
    """Generate a new access code for an employee."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Generate a new access code for the specified employee and send it by email."""
        employee_id = request.data.get("employee_id")
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response(
                {"error": "Empleado no encontrado."}, status=status.HTTP_404_NOT_FOUND
            )

        code_str = generate_code()
        code = AccessCode.objects.create(code=code_str, employee=employee)

        # Send email to employee with the access code and instructions
        subject = "Código de acceso para encuesta - Synergia"
        message = f"Hola {employee.name},\n\nTu código de acceso para la encuesta es: {code_str}\n\nImportante: El código se marcará como usado en cuanto hagas clic en 'Comenzar encuesta'. Si ocurre algún error y no puedes acceder, por favor contacta al administrador de tu empresa para que te genere un nuevo código.\n\nGracias por participar.\n\nEquipo Synergia"
        recipient_list = [employee.email]
        send_mail(
            subject,
            message,
            None,  # Usa DEFAULT_FROM_EMAIL
            recipient_list,
            fail_silently=False,
        )

        serializer = AccessCodeSerializer(code)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ListAccessCodesByEmployeeView(generics.ListAPIView):
    """List all access codes associated with a specific employee."""

    serializer_class = AccessCodeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Return access codes for the specified employee."""
        employee_id = self.kwargs["employee_id"]
        return AccessCode.objects.filter(employee__id=employee_id).order_by(
            "-created_at"
        )


class ValidateAccessCodeView(APIView):
    """Validate an access code to check if it is valid and has not been used."""

    permission_classes = [AllowAny]

    def post(self, request):
        """Validate the access code provided in the request data."""
        code = request.data.get("code")

        if not code:
            return Response(
                {"detail": "Código de acceso es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Buscar el código en la base de datos
            access_code = AccessCode.objects.get(code=code)

            if access_code.used:
                # Si el código ya ha sido usado
                return Response(
                    {"is_valid": False, "message": "El código ya ha sido utilizado."},
                    status=status.HTTP_200_OK,
                )

            # Si el código es válido, marcarlo como usado
            access_code.used = True
            access_code.save()

            # Responder con que el código es válido
            return Response({"is_valid": True}, status=status.HTTP_200_OK)

        except AccessCode.DoesNotExist:
            # Si el código no existe
            return Response(
                {"is_valid": False, "message": "Código no válido."},
                status=status.HTTP_200_OK,
            )


class SurveyStatisticsView(APIView):
    """View to get statistics about survey responses for a specific company."""

    permission_classes = [IsAuthenticated]

    def get(self, request, empresa_id):
        """Get statistics about survey responses for a specific company."""
        empleados = Employee.objects.filter(company_id=empresa_id)

        # Total de empleados registrados
        total_empleados = empleados.count()

        # Empleados que han respondido
        respuestas = SurveyResponse.objects.filter(employee__in=empleados)
        empleados_respondieron = respuestas.values("employee").distinct().count()

        # Total de respuestas
        total_respuestas = SurveyResponse.objects.filter(employee__in=empleados).count()

        # Promedio general de respuestas (en la escala de 1 a 5)
        promedio_respuestas = Answer.objects.filter(response__in=respuestas).aggregate(
            promedio=Avg("value")
        )["promedio"]

        # Promedio por pregunta (opcional)
        promedios_por_pregunta = (
            Answer.objects.filter(response__in=respuestas)
            .values("question__text")
            .annotate(promedio=Avg("value"))
            .order_by("question__text")
        )

        return Response(
            {
                "total_empleados": total_empleados,
                "empleados_que_respondieron": empleados_respondieron,
                "total_respuestas": total_respuestas,
                "promedio_general": round(promedio_respuestas or 0, 2),
                "promedios_por_pregunta": promedios_por_pregunta,
            }
        )


def _build_excel_report_context(business):
    responses_qs = (
        SurveyResponse.objects.filter(employee__company=business)
        .select_related("employee")
        .prefetch_related("answers__question__dimension")
        .order_by("submitted_at", "employee__name")
    )

    if not responses_qs.exists():
        return None

    responses = list(responses_qs)
    questions = list(Question.objects.select_related("dimension").order_by("order"))
    dimensions = list(Dimension.objects.prefetch_related("questions").all())
    answers = list(
        Answer.objects.filter(response__in=responses_qs).select_related(
            "question", "question__dimension", "response", "response__employee"
        )
    )

    answers_by_question = defaultdict(list)
    answers_by_dimension = defaultdict(list)
    answers_by_response = defaultdict(dict)

    for answer in answers:
        answers_by_question[answer.question_id].append(answer.value)
        answers_by_dimension[answer.question.dimension_id].append(answer.value)
        answers_by_response[answer.response_id][answer.question_id] = answer.value

    total_empleados = Employee.objects.filter(company=business).count()
    empleados_que_respondieron = len({response.employee_id for response in responses})
    total_respuestas = len(responses)
    promedio_general = (
        round(sum(answer.value for answer in answers) / len(answers), 2)
        if answers
        else 0
    )
    tasa_participacion = (
        round((empleados_que_respondieron / total_empleados) * 100, 2)
        if total_empleados
        else 0
    )

    dimension_rows = []
    for dimension in dimensions:
        question_list = list(dimension.questions.all().order_by("order"))
        promedio = answers_by_dimension[dimension.id]
        dimension_rows.append(
            {
                "dimension": dimension,
                "questions": question_list,
                "average": round(sum(promedio) / len(promedio), 2) if promedio else 0,
                "responses_count": len(promedio),
            }
        )

    dimension_rows.sort(
        key=lambda item: item["questions"][0].order if item["questions"] else 999
    )
    top_low_dimensions = sorted(dimension_rows, key=lambda item: item["average"])[:3]

    historical = list(
        Answer.objects.filter(response__in=responses_qs)
        .values("response__submitted_at__year", "response__submitted_at__month")
        .annotate(promedio=Avg("value"))
        .order_by("response__submitted_at__year", "response__submitted_at__month")
    )

    return {
        "business": business,
        "responses": responses,
        "questions": questions,
        "dimensions": dimension_rows,
        "answers_by_question": answers_by_question,
        "answers_by_response": answers_by_response,
        "total_empleados": total_empleados,
        "empleados_que_respondieron": empleados_que_respondieron,
        "total_respuestas": total_respuestas,
        "promedio_general": promedio_general,
        "tasa_participacion": tasa_participacion,
        "top_low_dimensions": top_low_dimensions,
        "historical": historical,
        "generated_at": timezone.localtime(),
    }


def _apply_sheet_base_style(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


def _apply_report_header(ws, title, subtitle, end_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color="334155")
    ws["A2"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")


def _write_section_title(ws, row, title, start_col=1, end_col=4):
    ws.merge_cells(
        start_row=row, start_column=start_col, end_row=row, end_column=end_col
    )
    cell = ws.cell(row=row, column=start_col)
    cell.value = title
    cell.font = Font(size=12, bold=True, color="0F172A")
    cell.fill = PatternFill("solid", fgColor="E2E8F0")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_excel_table(ws, start_row, end_row, start_col, end_col, table_name):
    if end_row <= start_row:
        return
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _auto_fit_columns(ws, extra_padding=2, max_width=45):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_length + extra_padding, max_width
        )


def _build_summary_sheet(workbook, context):
    ws = workbook.active
    ws.title = "Resumen Ejecutivo"
    _apply_sheet_base_style(ws)
    _apply_report_header(
        ws,
        "Reporte Ejecutivo de Clima Laboral",
        f"Empresa: {context['business'].name} | Generado: {context['generated_at'].strftime('%d/%m/%Y %H:%M')}",
        end_col=10,
    )

    _write_section_title(ws, 4, "Indicadores clave", start_col=1, end_col=2)
    kpi_headers = ["Indicador", "Valor"]
    for idx, header in enumerate(kpi_headers, start=1):
        ws.cell(row=5, column=idx, value=header)

    kpi_rows = [
        ["Empleados registrados", context["total_empleados"]],
        ["Empleados participantes", context["empleados_que_respondieron"]],
        ["Tasa de participación (%)", context["tasa_participacion"]],
        ["Respuestas registradas", context["total_respuestas"]],
        ["Promedio general", context["promedio_general"]],
    ]
    for row_offset, values in enumerate(kpi_rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_offset, column=col_idx, value=value)
    _apply_excel_table(ws, 5, 5 + len(kpi_rows), 1, 2, "SummaryKPI")

    _write_section_title(ws, 4, "Promedio por dimensión", start_col=4, end_col=5)
    ws.cell(row=5, column=4, value="Dimensión")
    ws.cell(row=5, column=5, value="Promedio")
    for idx, row in enumerate(context["dimensions"], start=6):
        ws.cell(row=idx, column=4, value=row["dimension"].name)
        ws.cell(row=idx, column=5, value=row["average"])
    _apply_excel_table(
        ws, 5, 5 + len(context["dimensions"]), 4, 5, "SummaryDimensionAverages"
    )

    _write_section_title(ws, 4, "Top 3 prioridades", start_col=7, end_col=8)
    ws.cell(row=5, column=7, value="Dimensión")
    ws.cell(row=5, column=8, value="Promedio")
    for idx, row in enumerate(context["top_low_dimensions"], start=6):
        ws.cell(row=idx, column=7, value=row["dimension"].name)
        ws.cell(row=idx, column=8, value=row["average"])
    _apply_excel_table(
        ws, 5, 5 + len(context["top_low_dimensions"]), 7, 8, "SummaryTopLowDimensions"
    )

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.height = 8
    bar_chart.width = 15
    bar_chart.title = "Promedio por dimensión"
    bar_chart.y_axis.title = "Promedio"
    bar_chart.x_axis.title = "Dimensión"
    data = Reference(ws, min_col=5, min_row=5, max_row=5 + len(context["dimensions"]))
    categories = Reference(
        ws, min_col=4, min_row=6, max_row=5 + len(context["dimensions"])
    )
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    ws.add_chart(bar_chart, "J4")

    history_start_row = max(16, 8 + len(context["dimensions"]))
    _write_section_title(ws, history_start_row, "Evolución histórica", end_col=6)
    ws.cell(row=history_start_row + 1, column=1, value="Periodo")
    ws.cell(row=history_start_row + 1, column=2, value="Promedio")
    for idx, item in enumerate(context["historical"], start=history_start_row + 2):
        ws.cell(
            row=idx,
            column=1,
            value=f"{item['response__submitted_at__month']:02d}/{item['response__submitted_at__year']}",
        )
        ws.cell(row=idx, column=2, value=round(item["promedio"] or 0, 2))
    if context["historical"]:
        _apply_excel_table(
            ws,
            history_start_row + 1,
            history_start_row + 1 + len(context["historical"]),
            1,
            2,
            "SummaryHistorical",
        )

        line_chart = LineChart()
        line_chart.style = 10
        line_chart.height = 8
        line_chart.width = 15
        line_chart.title = "Evolución del promedio general"
        line_chart.y_axis.title = "Promedio"
        line_chart.x_axis.title = "Periodo"
        line_data = Reference(
            ws,
            min_col=2,
            min_row=history_start_row + 1,
            max_row=history_start_row + 1 + len(context["historical"]),
        )
        line_categories = Reference(
            ws,
            min_col=1,
            min_row=history_start_row + 2,
            max_row=history_start_row + 1 + len(context["historical"]),
        )
        line_chart.add_data(line_data, titles_from_data=True)
        line_chart.set_categories(line_categories)
        ws.add_chart(line_chart, f"J{history_start_row}")

    _auto_fit_columns(ws)


def _build_dimension_sheet(workbook, context, dimension_index, dimension_row):
    dimension = dimension_row["dimension"]
    questions = dimension_row["questions"]
    ws = workbook.create_sheet(title=f"{dimension_index:02d}_{dimension.name}"[:31])
    _apply_sheet_base_style(ws)
    _apply_report_header(
        ws,
        f"Dimensión: {dimension.name}",
        f"Promedio: {dimension_row['average']} | Preguntas: {len(questions)} | Respuestas analizadas: {dimension_row['responses_count']}",
        end_col=10,
    )

    _write_section_title(ws, 4, "Promedio por pregunta", end_col=5)
    question_headers = ["Código", "Pregunta", "Promedio", "Respuestas"]
    for idx, header in enumerate(question_headers, start=1):
        ws.cell(row=5, column=idx, value=header)

    current_row = 6
    for question in questions:
        question_values = context["answers_by_question"][question.id]
        ws.cell(row=current_row, column=1, value=f"P{question.order}")
        ws.cell(row=current_row, column=2, value=question.text)
        ws.cell(
            row=current_row,
            column=3,
            value=round(sum(question_values) / len(question_values), 2)
            if question_values
            else 0,
        )
        ws.cell(row=current_row, column=4, value=len(question_values))
        current_row += 1

    if questions:
        _apply_excel_table(
            ws, 5, current_row - 1, 1, 4, f"DimensionQuestions{dimension_index}"
        )

        question_chart = BarChart()
        question_chart.type = "bar"
        question_chart.style = 10
        question_chart.height = 9
        question_chart.width = 12
        question_chart.title = f"Promedio por pregunta - {dimension.name}"
        question_chart.x_axis.title = "Promedio"
        question_chart.y_axis.title = "Pregunta"
        question_data = Reference(ws, min_col=3, min_row=5, max_row=current_row - 1)
        question_categories = Reference(
            ws, min_col=1, min_row=6, max_row=current_row - 1
        )
        question_chart.add_data(question_data, titles_from_data=True)
        question_chart.set_categories(question_categories)
        ws.add_chart(question_chart, "F5")

    detail_start_row = current_row + 3
    _write_section_title(
        ws,
        detail_start_row,
        "Detalle anónimo por respuesta",
        end_col=max(4, len(questions) + 3),
    )
    detail_headers = ["Fecha", "Registro"] + [
        f"P{question.order}" for question in questions
    ]
    for idx, header in enumerate(detail_headers, start=1):
        ws.cell(row=detail_start_row + 1, column=idx, value=header)

    detail_row = detail_start_row + 2
    for response in context["responses"]:
        response_answers = context["answers_by_response"][response.id]
        if not any(question.id in response_answers for question in questions):
            continue
        ws.cell(
            row=detail_row, column=1, value=response.submitted_at.strftime("%d/%m/%Y")
        )
        ws.cell(row=detail_row, column=2, value=f"R-{response.id}")
        for q_index, question in enumerate(questions, start=3):
            ws.cell(
                row=detail_row,
                column=q_index,
                value=response_answers.get(question.id, ""),
            )
        detail_row += 1

    if detail_row > detail_start_row + 2:
        _apply_excel_table(
            ws,
            detail_start_row + 1,
            detail_row - 1,
            1,
            len(detail_headers),
            f"DimensionAnonymousDetail{dimension_index}",
        )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    _auto_fit_columns(ws, max_width=30)
    ws.column_dimensions["B"].width = 60


def _build_identified_detail_sheet(workbook, context):
    ws = workbook.create_sheet(title="Detalle Identificado")
    _apply_sheet_base_style(ws)
    _apply_report_header(
        ws,
        "Detalle Identificado de Respuestas",
        "Hoja interna para seguimiento por empleado. Incluye identidad y respuestas completas.",
        end_col=8 + len(context["questions"]),
    )

    headers = ["Fecha", "Empleado", "Correo", "Departamento", "Cargo"] + [
        f"P{question.order}" for question in context["questions"]
    ]
    header_row = 5
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=header)

    data_row = 6
    for response in context["responses"]:
        employee = response.employee
        answer_map = context["answers_by_response"][response.id]
        row_values = [
            response.submitted_at.strftime("%d/%m/%Y"),
            employee.name,
            employee.email,
            employee.department,
            employee.position,
        ]
        row_values.extend(
            answer_map.get(question.id, "") for question in context["questions"]
        )
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=data_row, column=col_idx, value=value)
        data_row += 1

    if data_row > 6:
        _apply_excel_table(
            ws, header_row, data_row - 1, 1, len(headers), "IdentifiedDetailTable"
        )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    _auto_fit_columns(ws, max_width=16)
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20


class ExportSurveyDataView(APIView):
    """View to export survey responses as a CSV file."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Export survey responses as a CSV file."""
        business = request.user.business
        questions = Question.objects.all().order_by("order")
        responses = (
            SurveyResponse.objects.filter(employee__company=business)
            .select_related("employee")
            .prefetch_related("answers__question")
            .order_by("employee__name", "-submitted_at")
        )

        if not responses.exists():
            return HttpResponse(
                "No hay respuestas disponibles para exportar.", status=400
            )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            'attachment; filename="respuestas_encuesta.csv"'
        )
        writer = csv.writer(response)

        headers = ["Fecha"]
        headers.extend([f"Pregunta {q.order}" for q in questions])
        writer.writerow(headers)

        for resp in responses:
            answers_dict = {
                answer.question.order: answer.value for answer in resp.answers.all()
            }
            row = [resp.submitted_at.strftime("%d/%m/%Y")]
            row.extend([answers_dict.get(q.order, "") for q in questions])
            writer.writerow(row)

        return response


class ExportSurveyExcelView(APIView):
    """Export survey data as a professional Excel workbook."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        business = request.user.business
        context = _build_excel_report_context(business)

        if not context:
            return HttpResponse(
                "No hay respuestas disponibles para exportar.", status=400
            )

        workbook = Workbook()
        _build_summary_sheet(workbook, context)

        for index, dimension_row in enumerate(context["dimensions"], start=1):
            _build_dimension_sheet(workbook, context, index, dimension_row)

        _build_identified_detail_sheet(workbook, context)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"reporte_clima_{slugify(business.name) or 'empresa'}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def _build_ai_summary_context(report_context):
    question_rows = []
    for question in report_context["questions"]:
        values = report_context["answers_by_question"].get(question.id, [])
        promedio = round(sum(values) / len(values), 2) if values else 0
        question_rows.append(
            {
                "dimension": question.dimension.name,
                "order": question.order,
                "text": question.text,
                "average": promedio,
                "responses": len(values),
            }
        )

    low_questions = sorted(question_rows, key=lambda item: item["average"])[:5]
    high_questions = sorted(
        question_rows, key=lambda item: item["average"], reverse=True
    )[:5]
    low_dimensions = [
        {"dimension": item["dimension"].name, "average": item["average"]}
        for item in sorted(
            report_context["dimensions"], key=lambda item: item["average"]
        )[:3]
    ]
    high_dimensions = [
        {"dimension": item["dimension"].name, "average": item["average"]}
        for item in sorted(
            report_context["dimensions"], key=lambda item: item["average"], reverse=True
        )[:3]
    ]

    historical = [
        {
            "period": f"{item['response__submitted_at__month']:02d}/{item['response__submitted_at__year']}",
            "average": round(item["promedio"] or 0, 2),
        }
        for item in report_context["historical"]
    ]

    return {
        "empresa": report_context["business"].name,
        "fecha_generacion": report_context["generated_at"].strftime("%d/%m/%Y %H:%M"),
        "total_empleados": report_context["total_empleados"],
        "empleados_participantes": report_context["empleados_que_respondieron"],
        "tasa_participacion": report_context["tasa_participacion"],
        "total_respuestas": report_context["total_respuestas"],
        "promedio_general": report_context["promedio_general"],
        "dimensiones_mas_bajas": low_dimensions,
        "dimensiones_mas_altas": high_dimensions,
        "preguntas_mas_bajas": low_questions,
        "preguntas_mas_altas": high_questions,
        "historico": historical,
    }


def _extract_json_from_gemini_text(raw_text):
    raw_text = raw_text.strip()
    if raw_text.startswith("```"):
        raw_text = raw_text.strip("`")
        if raw_text.startswith("json"):
            raw_text = raw_text[4:].strip()
    start = raw_text.find("{")
    end = raw_text.rfind("}")
    if start != -1 and end != -1 and end > start:
        raw_text = raw_text[start : end + 1]
    return json.loads(raw_text)


def _priority_from_average(avg):
    if avg < 2.5:
        return "Alta"
    if avg < 3.5:
        return "Media"
    return "Baja"


def _build_local_ai_summary(ai_context):
    low_dimensions = ai_context.get("dimensiones_mas_bajas", [])
    high_dimensions = ai_context.get("dimensiones_mas_altas", [])
    low_questions = ai_context.get("preguntas_mas_bajas", [])
    overall = ai_context.get("promedio_general", 0)
    participation = ai_context.get("tasa_participacion", 0)

    if overall < 2.5:
        diagnostico = "El clima laboral general presenta señales de alerta y requiere intervención prioritaria."
    elif overall < 3.5:
        diagnostico = "El clima laboral es aceptable, pero existen áreas claras de mejora que conviene atender a corto plazo."
    else:
        diagnostico = "El clima laboral es favorable en términos generales, aunque aún hay oportunidades de fortalecimiento puntual."

    fortalezas = [
        f"La dimensión {item['dimension']} muestra un desempeño relativamente fuerte con promedio {item['average']}."
        for item in high_dimensions[:3]
        if item.get("average", 0) > 0
    ]

    areas_criticas = [
        f"La dimensión {item['dimension']} requiere atención porque su promedio es {item['average']}."
        for item in low_dimensions[:3]
        if item.get("average", 0) > 0
    ]

    preguntas_clave = [
        {
            "dimension": item["dimension"],
            "pregunta": item["text"],
            "promedio": item["average"],
            "interpretacion": (
                "Este ítem evidencia una percepción débil y puede estar afectando negativamente la experiencia laboral del equipo."
                if item.get("average", 0) < 2.5
                else "Este ítem sugiere una percepción intermedia que aún puede fortalecerse con acciones específicas."
            ),
        }
        for item in low_questions[:4]
    ]

    recomendaciones = []
    for item in low_dimensions[:3]:
        recomendaciones.append(
            {
                "titulo": f"Mejorar {item['dimension']}",
                "detalle": f"Diseñar acciones concretas para reforzar la dimensión {item['dimension']}, revisar prácticas actuales y hacer seguimiento mensual de avances.",
                "prioridad": _priority_from_average(item.get("average", 0)),
            }
        )

    if participation < 70:
        recomendaciones.append(
            {
                "titulo": "Incrementar participación",
                "detalle": "Reforzar la comunicación interna y facilitar el acceso a la encuesta para lograr una muestra más representativa.",
                "prioridad": "Media",
            }
        )

    plan_accion = [
        {
            "plazo": "Corto plazo",
            "accion": "Socializar resultados con líderes y priorizar las dimensiones más bajas.",
            "objetivo": "Alinear decisiones rápidas basadas en evidencia.",
        },
        {
            "plazo": "Mediano plazo",
            "accion": "Implementar intervenciones específicas sobre reconocimiento, comunicación o liderazgo según la dimensión crítica detectada.",
            "objetivo": "Mejorar la percepción de los equipos en los factores con peor desempeño.",
        },
        {
            "plazo": "Largo plazo",
            "accion": "Comparar nuevas mediciones con esta línea base y consolidar una estrategia de mejora continua.",
            "objetivo": "Sostener mejoras reales en el clima laboral.",
        },
    ]

    riesgos = []
    if participation < 50:
        riesgos.append(
            "La baja participación puede limitar la representatividad del diagnóstico."
        )
    if low_dimensions and low_dimensions[0].get("average", 0) < 2.5:
        riesgos.append(
            f"La dimensión {low_dimensions[0]['dimension']} presenta un nivel crítico que podría impactar el compromiso, la motivación o la convivencia laboral."
        )
    if not riesgos:
        riesgos.append(
            "No se detectan riesgos severos inmediatos, pero conviene mantener seguimiento periódico."
        )

    return {
        "resumen_ejecutivo": f"Se analizó la información agregada de la empresa {ai_context.get('empresa')}. El promedio general es {overall} sobre 5 y la tasa de participación alcanza {participation}%. Las principales oportunidades de mejora se concentran en las dimensiones con menor puntaje y en los ítems críticos detectados.",
        "diagnostico_general": diagnostico,
        "fortalezas": fortalezas,
        "areas_criticas": areas_criticas,
        "preguntas_clave": preguntas_clave,
        "recomendaciones": recomendaciones,
        "plan_accion": plan_accion,
        "riesgos": riesgos,
        "conclusion": "El análisis sugiere que la organización cuenta con una base útil para actuar de forma focalizada. La prioridad debe centrarse en las dimensiones más débiles, acompañada de seguimiento periódico y comunicación clara con los equipos.",
        "fuente": "fallback-local",
        "modelo": "reglas-internas",
        "warning": "Gemini no pudo responder por límite de cuota o disponibilidad. Se generó un resumen local automático como respaldo.",
    }


class SurveyAISummaryView(APIView):
    """Generate an AI-powered survey summary using Gemini."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        api_key = getattr(django_settings, "GEMINI_API_KEY", "")
        if not api_key:
            return Response(
                {"error": "No se ha configurado GEMINI_API_KEY en settings.py."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        business = request.user.business
        report_context = _build_excel_report_context(business)
        if not report_context:
            return Response(
                {
                    "error": "No hay respuestas suficientes para generar el resumen inteligente."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        ai_context = _build_ai_summary_context(report_context)
        model = getattr(django_settings, "GEMINI_MODEL", "gemini-2.0-flash")
        prompt = f"""
Eres un analista senior de clima laboral. Debes analizar los datos agregados de una empresa y devolver exclusivamente un JSON válido, sin markdown, sin comentarios y sin texto adicional.

Responde en español profesional, con lenguaje claro y accionable.

Debes devolver exactamente esta estructura JSON:
{{
  "resumen_ejecutivo": "string",
  "diagnostico_general": "string",
  "fortalezas": ["string"],
  "areas_criticas": ["string"],
  "preguntas_clave": [
    {{
      "dimension": "string",
      "pregunta": "string",
      "promedio": number,
      "interpretacion": "string"
    }}
  ],
  "recomendaciones": [
    {{
      "titulo": "string",
      "detalle": "string",
      "prioridad": "Alta|Media|Baja"
    }}
  ],
  "plan_accion": [
    {{
      "plazo": "Corto plazo|Mediano plazo|Largo plazo",
      "accion": "string",
      "objetivo": "string"
    }}
  ],
  "riesgos": ["string"],
  "conclusion": "string"
}}

Usa solo la información disponible. No inventes dimensiones, preguntas, métricas ni datos.

Datos agregados:
{json.dumps(ai_context, ensure_ascii=False, indent=2)}
"""

        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
        payload = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.4,
                "responseMimeType": "application/json",
            },
        }

        try:
            gemini_response = requests.post(url, json=payload, timeout=40)
            gemini_response.raise_for_status()
            data = gemini_response.json()
            raw_text = data["candidates"][0]["content"]["parts"][0]["text"]
            parsed = _extract_json_from_gemini_text(raw_text)
            parsed["fuente"] = "gemini"
            parsed["modelo"] = model
            return Response(parsed)
        except requests.HTTPError as exc:
            status_code = exc.response.status_code if exc.response is not None else None
            if status_code == 429:
                fallback_summary = _build_local_ai_summary(ai_context)
                return Response(fallback_summary)
            return Response(
                {"error": f"No fue posible comunicarse con Gemini: {str(exc)}"},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except requests.RequestException as exc:
            return Response(
                {"error": f"No fue posible comunicarse con Gemini: {str(exc)}"},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except (KeyError, IndexError, json.JSONDecodeError):
            fallback_summary = _build_local_ai_summary(ai_context)
            fallback_summary["warning"] = (
                "Gemini respondió, pero su salida no pudo procesarse. Se generó un resumen local automático como respaldo."
            )
            return Response(fallback_summary)


class SurveyAnalysisView(APIView):
    """View to analyze survey responses and provide descriptive statistics solo para la empresa autenticada."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Analyze survey responses and provide descriptive statistics solo para la empresa autenticada."""
        business = request.user.business
        empleados = Employee.objects.filter(company=business)
        respuestas = SurveyResponse.objects.filter(employee__in=empleados)
        dimensiones = Dimension.objects.all()
        resultados = []

        for dimension in dimensiones:
            # Filtrar solo respuestas de empleados de la empresa autenticada y de la dimensión
            answers_dim = Answer.objects.filter(
                response__in=respuestas, question__dimension=dimension
            )
            promedio = answers_dim.aggregate(promedio_respuesta=Avg("value"))[
                "promedio_respuesta"
            ]
            desviacion_estandar = answers_dim.aggregate(
                desviacion_estandar=StdDev("value")
            )["desviacion_estandar"]

            # Pregunta con mayor variabilidad dentro de la empresa
            pregunta_mas_variable = (
                answers_dim.values("question")
                .annotate(desviacion=StdDev("value"))
                .order_by("-desviacion")
                .first()
            )

            texto_pregunta_mas_variable = None
            if pregunta_mas_variable and pregunta_mas_variable["question"]:
                try:
                    texto_pregunta_mas_variable = Question.objects.get(
                        id=pregunta_mas_variable["question"]
                    ).text
                except Question.DoesNotExist:
                    texto_pregunta_mas_variable = None

            resultados.append(
                {
                    "dimensión": dimension.name,
                    "promedio": promedio,
                    "desviacion_estandar": desviacion_estandar,
                    "pregunta_mas_variable": pregunta_mas_variable["question"]
                    if pregunta_mas_variable
                    else None,
                    "texto_pregunta_mas_variable": texto_pregunta_mas_variable,
                }
            )

        return Response(resultados)


class SurveyRecommendationsView(APIView):
    """Genera recomendaciones automáticas y datos para gráficas para la empresa autenticada."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        business = request.user.business
        empleados = Employee.objects.filter(company=business)
        respuestas = SurveyResponse.objects.filter(employee__in=empleados)
        dimensiones = Dimension.objects.all()

        # Promedio general solo de respuestas de la empresa
        promedio_general = Answer.objects.filter(response__in=respuestas).aggregate(
            prom=Avg("value")
        )["prom"]
        promedio_general = round(promedio_general or 0, 2)

        # Promedios por dimensión solo de la empresa
        promedios_dim = []
        recomendaciones_dim = []
        radar_labels = []
        radar_data = []
        for dim in dimensiones:
            prom = Answer.objects.filter(
                response__in=respuestas, question__dimension=dim
            ).aggregate(prom=Avg("value"))["prom"]
            prom = round(prom or 0, 2)
            promedios_dim.append({"dimension": dim.name, "promedio": prom})
            radar_labels.append(dim.name)
            radar_data.append(prom)
            # Reglas automáticas para recomendaciones
            if prom < 2.5:
                recomendaciones_dim.append(
                    f"La dimensión '{dim.name}' tiene un puntaje bajo ({prom}). Se recomienda realizar talleres, capacitaciones o reuniones para mejorar este aspecto."
                )
            elif prom < 3.5:
                recomendaciones_dim.append(
                    f"La dimensión '{dim.name}' es mejorable (puntaje: {prom}). Considere encuestas internas o focus groups para identificar oportunidades de mejora."
                )
            else:
                recomendaciones_dim.append(
                    f"La dimensión '{dim.name}' está bien valorada (puntaje: {prom}). Mantenga las buenas prácticas actuales."
                )

        # Recomendación general
        if promedio_general < 2.5:
            recomendacion_general = "El clima laboral general es bajo. Se recomienda una intervención integral, comunicación directa con los empleados y revisión de procesos internos."
        elif promedio_general < 3.5:
            recomendacion_general = "El clima laboral es aceptable pero mejorable. Considere reforzar la motivación, reconocimiento y espacios de escucha activa."
        else:
            recomendacion_general = "El clima laboral es positivo. Continúe con las estrategias actuales y refuerce la cultura organizacional."

        # Histórico de promedios generales (por mes) solo de la empresa
        historico = (
            Answer.objects.filter(response__in=respuestas)
            .values("response__submitted_at__month", "response__submitted_at__year")
            .annotate(prom=Avg("value"))
            .order_by("response__submitted_at__year", "response__submitted_at__month")
        )
        historico_labels = [
            f"{h['response__submitted_at__month']}/{h['response__submitted_at__year']}"
            for h in historico
        ]
        historico_data = [round(h["prom"] or 0, 2) for h in historico]

        return Response(
            {
                "promedio_general": promedio_general,
                "promedios_dimensiones": promedios_dim,
                "recomendacion_general": recomendacion_general,
                "recomendaciones_dimensiones": recomendaciones_dim,
                "radar_labels": radar_labels,
                "radar_data": radar_data,
                "historico_labels": historico_labels,
                "historico_data": historico_data,
            }
        )


class ExportSurveyAllDataView(APIView):
    """Exporta los resultados de encuestas de todas las empresas en CSV o Excel, con hoja/archivo por empresa y fila de promedios."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Solo admins pueden exportar todo
        if not request.user.is_superuser:
            return Response({"error": "No autorizado"}, status=403)
        from employees.models import Company

        questions = list(Question.objects.all().order_by("order"))
        question_headers = [f"Pregunta {q.order}" for q in questions]
        empresas = Company.objects.all()
        temp_files = []
        if request.query_params.get("format", "csv") == "excel":
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                writer = pd.ExcelWriter(tmp.name, engine="openpyxl")
                for empresa in empresas:
                    empleados = Employee.objects.filter(company=empresa)
                    responses = SurveyResponse.objects.filter(
                        employee__in=empleados
                    ).order_by("submitted_at")
                    data = []
                    for resp in responses:
                        answers_dict = {
                            a.question_id: a.value for a in resp.answers.all()
                        }
                        row = [resp.submitted_at.strftime("%Y-%m-%d")]
                        for q in questions:
                            row.append(answers_dict.get(q.id, ""))
                        data.append(row)
                    df = pd.DataFrame(data, columns=["Fecha"] + question_headers)
                    # Fila de promedios (solo columnas de preguntas)
                    if not df.empty:
                        promedios = ["Promedio"] + [
                            df[q].astype(float).mean(skipna=True)
                            for q in question_headers
                        ]
                        df.loc[len(df)] = promedios
                    # Nombre limpio para la hoja
                    sheet_name = slugify(empresa.name)[:31] or f"empresa_{empresa.id}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.close()
                tmp.seek(0)
                response = FileResponse(
                    open(tmp.name, "rb"),
                    as_attachment=True,
                    filename="synergia_export.xlsx",
                )
                temp_files.append(tmp.name)
                return response
        else:  # CSV
            # Un archivo zip con un CSV por empresa
            import zipfile

            with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as zip_tmp:
                with zipfile.ZipFile(zip_tmp, "w") as zipf:
                    for empresa in empresas:
                        empleados = Employee.objects.filter(company=empresa)
                        responses = SurveyResponse.objects.filter(
                            employee__in=empleados
                        ).order_by("submitted_at")
                        data = []
                        for resp in responses:
                            answers_dict = {
                                a.question_id: a.value for a in resp.answers.all()
                            }
                            row = [resp.submitted_at.strftime("%Y-%m-%d")]
                            for q in questions:
                                row.append(answers_dict.get(q.id, ""))
                            data.append(row)
                        df = pd.DataFrame(data, columns=["Fecha"] + question_headers)
                        # Fila de promedios
                        if not df.empty:
                            promedios = ["Promedio"] + [
                                df[q].astype(float).mean(skipna=True)
                                for q in question_headers
                            ]
                            df.loc[len(df)] = promedios
                        # Guardar CSV temporal
                        csv_tmp = tempfile.NamedTemporaryFile(
                            delete=False, suffix=".csv"
                        )
                        df.to_csv(csv_tmp.name, index=False, encoding="utf-8-sig")
                        csv_tmp.close()
                        # Nombre limpio para el archivo
                        csv_name = slugify(empresa.name) or f"empresa_{empresa.id}"
                        zipf.write(csv_tmp.name, arcname=f"{csv_name}.csv")
                        temp_files.append(csv_tmp.name)
                zip_tmp.seek(0)
                response = FileResponse(
                    open(zip_tmp.name, "rb"),
                    as_attachment=True,
                    filename="synergia_export.zip",
                )
                temp_files.append(zip_tmp.name)
                return response
        # Limpieza de archivos temporales (opcional, depende del servidor)
        for f in temp_files:
            try:
                os.remove(f)
            except Exception:
                pass
